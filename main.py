"""
IB Pitch Multi-Agent System — Analyst Pipeline
===============================================
Entry point. Run with:
  python main.py
Output: Excel file in outputs/ folder
"""

from state import DealContext, AnalystState
from runner import SubAgentRunner
from pipeline import build_analyst_graph
from human_gate import human_gate


def get_user_inputs() -> dict:
    """Prompt user for company details in the terminal."""
    print("\n" + "=" * 60)
    print("  IB PITCH ANALYST SYSTEM")
    print("=" * 60)

    company = input("\nCompany Name       : ").strip()
    ticker  = input("Ticker Symbol      : ").strip()

    print("\nSector options:")
    sectors = [
        "Power / Utilities",
        "Oil & Gas / Conglomerate",
        "Banking & Financial Services",
        "IT Services",
        "FMCG",
        "Telecommunications",
    ]
    for i, s in enumerate(sectors, 1):
        print(f"  {i}. {s}")
    sec_idx = input("Choose sector (1-6): ").strip()
    sector  = sectors[int(sec_idx) - 1] if sec_idx.isdigit() and 1 <= int(sec_idx) <= 6 else sec_idx

    print("\nGeography options:  India / USA / UK / Singapore / UAE")
    geography = input("Geography          : ").strip() or "India"

    print("\nTransaction Type:   1. Buy  2. Sell  3. IPO  4. Merger  5. Acquisition")
    txn_map = {"1": "buy-side_advisory", "2": "sell-side_advisory", "3": "ipo", "4": "merger", "5": "acquisition"}
    txn_idx = input("Choose type (1-5)  : ").strip()
    transaction_type = txn_map.get(txn_idx, "sell-side_advisory")

    print("\nEnter business segments (comma-separated).")
    print("Example: Thermal Generation, Renewable Energy, Distribution")
    seg_input = input("Segments           : ").strip()
    segments  = [s.strip() for s in seg_input.split(",") if s.strip()]

    print("\nOptional: Path to a PDF file (quarterly results / annual report).")
    print("Press Enter to skip.")
    pdf_path = input("PDF path           : ").strip()
    pdf_text = ""
    if pdf_path:
        try:
            import pypdf
            reader = pypdf.PdfReader(pdf_path)
            for page in reader.pages:
                pdf_text += page.extract_text() + "\n"
            pdf_text = pdf_text[:8000]
            print(f"  ✅ PDF loaded — {len(pdf_text)} characters extracted")
        except Exception as e:
            print(f"  ⚠️  Could not read PDF: {e}")

    return {
        "company":          company,
        "ticker":           ticker,
        "sector":           sector,
        "geography":        geography,
        "transaction_type": transaction_type,
        "segments":         segments,
        "pdf_text":         pdf_text,
    }


def run_analyst_pipeline(inputs: dict = None):
    """Execute the full Analyst Agent pipeline."""

    if inputs is None:
        inputs = get_user_inputs()

    deal_context: DealContext = {
        "target_name":      inputs["company"],
        "target_ticker":    inputs["ticker"],
        "sector":           inputs["sector"],
        "segments":         inputs["segments"],
        "transaction_type": inputs["transaction_type"],
        "listed":           True,
        "geography":        inputs["geography"],
        "pdf_data":         inputs.get("pdf_text", ""),
    }

    analyst_manifest = {
        "priority": "high",
        "deadline": "T+48h",
        "iteration_budget": 3,
        "quality_thresholds": {
            "balance_check": True,
            "source_citation": "every_assumption",
            "segment_reconciliation": True,
        },
    }

    initial_state: AnalystState = {
        "deal_context": deal_context,
        "analyst_manifest": analyst_manifest,
        "raw_data": None,
        "financial_model": None,
        "valuation": None,
        "benchmarking": None,
        "analyst_package": None,
        "consistency_report": None,
        "human_feedback": None,
        "status": "running",
        "errors": [],
    }

    print("\n" + "=" * 60)
    print("IB PITCH AGENT SYSTEM — Analyst Pipeline")
    print(f"  Target  : {deal_context['target_name']}")
    print(f"  Sector  : {deal_context['sector']}")
    print(f"  Segments: {', '.join(deal_context['segments'])}")
    print("=" * 60)

    runner = SubAgentRunner()
    compiled_graph = build_analyst_graph(runner)

    if compiled_graph is not None:
        config = {"configurable": {"thread_id": "tata-power-pitch-001"}}
        final_state = compiled_graph.invoke(initial_state, config)
    else:
        print("\nRunning sequential fallback (no LangGraph)...")
        import agents.data_sourcing     as data_sourcing
        import agents.financial_modeler as financial_modeler
        import agents.benchmarking      as benchmarking
        import agents.valuation         as valuation
        import agents.assembly          as assembly

        state = dict(initial_state)
        state.update(data_sourcing.run(state, runner))
        state.update(financial_modeler.run(state, runner))
        state.update(benchmarking.run(state, runner))
        state.update(valuation.run(state, runner))
        state.update(assembly.run(state, runner))
        final_state = state

    feedback = human_gate(
        final_state.get("analyst_package", {}),
        final_state.get("consistency_report", {}),
    )

    runner.print_cost_summary()

    return final_state, feedback


# ─────────────────────────────────────────────────────────────
# Excel writer
# ─────────────────────────────────────────────────────────────

def write_excel(final_state: dict, feedback: dict, output_path: str):
    """Write all pipeline outputs into a formatted Excel workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        return

    wb = Workbook()

    # ── Colour palette ──────────────────────────────────────────
    DARK_BLUE  = "003366"
    LIGHT_BLUE = "D9E1F2"
    GOLD       = "C9A84C"
    WHITE      = "FFFFFF"
    LIGHT_GREY = "F2F2F2"

    def header_font(color=WHITE):
        return Font(name="Calibri", bold=True, color=color, size=11)

    def normal_font():
        return Font(name="Calibri", size=10)

    def title_font():
        return Font(name="Calibri", bold=True, color=DARK_BLUE, size=14)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def set_col_width(ws, col, width):
        ws.column_dimensions[get_column_letter(col)].width = width

    def write_section_title(ws, row, col, text, colspan=6):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=12)
        cell.fill = fill(LIGHT_BLUE)
        cell.alignment = left()
        cell.border = thin_border()
        if colspan > 1:
            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row, end_column=col + colspan - 1
            )

    def write_header_row(ws, row, headers, start_col=1):
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=start_col + i, value=h)
            cell.font = header_font()
            cell.fill = fill(DARK_BLUE)
            cell.alignment = center()
            cell.border = thin_border()

    def write_data_row(ws, row, values, start_col=1, shade=False):
        bg = LIGHT_GREY if shade else WHITE
        for i, v in enumerate(values):
            cell = ws.cell(row=row, column=start_col + i, value=v)
            cell.font = normal_font()
            cell.fill = fill(bg)
            cell.alignment = left()
            cell.border = thin_border()

    def flatten(obj, prefix=""):
        """Flatten nested dict to list of (key, value) pairs."""
        items = []
        if isinstance(obj, dict):
            for k, v in obj.items():
                full_key = f"{prefix}.{k}" if prefix else k
                if isinstance(v, (dict, list)):
                    items.extend(flatten(v, full_key))
                else:
                    items.append((full_key, str(v) if v is not None else ""))
        elif isinstance(obj, list):
            for i, v in enumerate(obj):
                items.extend(flatten(v, f"{prefix}[{i}]"))
        else:
            items.append((prefix, str(obj) if obj is not None else ""))
        return items

    # ══════════════════════════════════════════════════════════════
    # Sheet 1: SUMMARY
    # ══════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 20

    # Title banner
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "TATA POWER — IB PITCH ANALYST PACKAGE"
    t.font = Font(name="Calibri", bold=True, color=WHITE, size=16)
    t.fill = fill(DARK_BLUE)
    t.alignment = center()

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = "Goldman Sachs India IBD | Sell-Side Advisory | Confidential"
    sub.font = Font(name="Calibri", italic=True, color=DARK_BLUE, size=10)
    sub.fill = fill(LIGHT_BLUE)
    sub.alignment = center()

    ctx = final_state.get("deal_context", {})
    row = 4
    write_section_title(ws, row, 1, "Deal Context", 3)
    row += 1
    for label, val in [
        ("Company",          ctx.get("target_name", "")),
        ("Ticker",           ctx.get("target_ticker", "")),
        ("Sector",           ctx.get("sector", "")),
        ("Geography",        ctx.get("geography", "")),
        ("Transaction Type", ctx.get("transaction_type", "")),
        ("Segments",         ", ".join(ctx.get("segments", []))),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(row=row, column=1).fill = fill(LIGHT_GREY)
        ws.cell(row=row, column=2, value=val).font = normal_font()
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1

    row += 1
    write_section_title(ws, row, 1, "Pipeline Status", 3)
    row += 1
    cr = final_state.get("consistency_report") or {}
    for label, val in [
        ("Human Gate Decision", feedback.get("decision", "").upper()),
        ("Consistency Status",  cr.get("status", "N/A").upper()),
        ("Approved Artifacts",  ", ".join(feedback.get("approved_artifacts", []))),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(row=row, column=1).fill = fill(LIGHT_GREY)
        ws.cell(row=row, column=2, value=val).font = normal_font()
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1

    for col, width in [(1, 25), (2, 30), (3, 20), (4, 20), (5, 20), (6, 20)]:
        set_col_width(ws, col, width)

    # ══════════════════════════════════════════════════════════════
    # Sheet 2: FINANCIAL MODEL
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Financial Model")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    t2 = ws2["A1"]
    t2.value = "FINANCIAL MODEL — 5-YEAR PROJECTIONS"
    t2.font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    t2.fill = fill(DARK_BLUE)
    t2.alignment = center()
    ws2.row_dimensions[1].height = 35

    fm = final_state.get("financial_model") or {}

    if fm.get("parse_error"):
        ws2.cell(row=3, column=1, value="Raw Output (parse error — see below):").font = Font(bold=True, name="Calibri")
        raw = fm.get("raw_output", "")
        lines = str(raw).split("\n")
        for i, line in enumerate(lines[:200]):
            ws2.cell(row=4 + i, column=1, value=line).font = normal_font()
    else:
        row = 3
        # Projections table
        projections = fm.get("projections", {})
        if projections:
            write_section_title(ws2, row, 1, "Revenue & EBITDA Projections (INR Cr)", 7)
            row += 1
            years = ["FY25", "FY26", "FY27", "FY28", "FY29"]
            write_header_row(ws2, row, ["Metric"] + years + ["CAGR"], 1)
            row += 1
            for metric_name, values in projections.items():
                if isinstance(values, list):
                    data_row = [metric_name.replace("_", " ").title()] + [str(v) for v in values[:5]]
                    write_data_row(ws2, row, data_row, shade=(row % 2 == 0))
                elif isinstance(values, dict):
                    data_row = [metric_name.replace("_", " ").title()] + [str(values.get(y, "")) for y in years]
                    write_data_row(ws2, row, data_row, shade=(row % 2 == 0))
                row += 1

        row += 1
        # Assumptions table
        assumptions = fm.get("assumptions", {})
        if assumptions:
            write_section_title(ws2, row, 1, "Key Assumptions", 7)
            row += 1
            write_header_row(ws2, row, ["Assumption Category", "Value / Description", "Rationale"], 1)
            ws2.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
            row += 1
            flat = flatten(assumptions)
            for i, (k, v) in enumerate(flat):
                ws2.cell(row=row, column=1, value=k).font = normal_font()
                ws2.cell(row=row, column=2, value=v).font = normal_font()
                if i % 2 == 0:
                    ws2.cell(row=row, column=1).fill = fill(LIGHT_GREY)
                    ws2.cell(row=row, column=2).fill = fill(LIGHT_GREY)
                row += 1

        row += 1
        # All remaining keys
        skip = {"projections", "assumptions"}
        remaining = {k: v for k, v in fm.items() if k not in skip}
        if remaining:
            write_section_title(ws2, row, 1, "Additional Model Data", 7)
            row += 1
            write_header_row(ws2, row, ["Field", "Value"], 1)
            row += 1
            for k, v in flatten(remaining):
                write_data_row(ws2, row, [k, v], shade=(row % 2 == 0))
                row += 1

    for col, w in [(1, 30), (2, 18), (3, 18), (4, 18), (5, 18), (6, 18), (7, 15)]:
        set_col_width(ws2, col, w)

    # ── Formula Reference Table (Financial Model) ───────────────
    row += 2
    write_section_title(ws2, row, 1, "Formula Reference", 3)
    row += 1
    write_header_row(ws2, row, ["Output", "Formula", "Notes"], 1)
    row += 1
    fm_formulas = [
        ("Revenue (FY+1)",      "Revenue_t = Revenue_(t-1) × (1 + Growth Rate)",         "CAGR-based projection"),
        ("EBITDA",              "EBITDA = Revenue × EBITDA Margin %",                     "Margin applied to revenue"),
        ("EBIT",                "EBIT = EBITDA − Depreciation",                           "D&A subtracted"),
        ("PAT (Net Profit)",    "PAT = EBIT × (1 − Tax Rate)",                            "Effective tax rate applied"),
        ("Free Cash Flow",      "FCF = EBITDA − Tax − Capex − ΔWorking Capital",          "Unlevered FCF"),
        ("Net Working Capital", "NWC = Current Assets − Current Liabilities",              "Balance sheet derived"),
        ("Net Debt",            "Net Debt = Total Debt − Cash & Equivalents",              "Used in EV bridge"),
        ("Balance Check",       "Total Assets = Total Liabilities + Shareholders Equity",  "Must balance"),
    ]
    for i, (out, formula, note) in enumerate(fm_formulas):
        write_data_row(ws2, row, [out, formula, note], shade=(i % 2 == 0))
        row += 1

    # ══════════════════════════════════════════════════════════════
    # Sheet 3: VALUATION
    # ══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Valuation")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:F1")
    t3 = ws3["A1"]
    t3.value = "VALUATION SUMMARY — DCF | SOTP | TRADING COMPS | PRECEDENT TRANSACTIONS"
    t3.font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    t3.fill = fill(DARK_BLUE)
    t3.alignment = center()
    ws3.row_dimensions[1].height = 35

    val = final_state.get("valuation") or {}

    if val.get("parse_error"):
        ws3.cell(row=3, column=1, value="Raw Output:").font = Font(bold=True, name="Calibri")
        for i, line in enumerate(str(val.get("raw_output", "")).split("\n")[:200]):
            ws3.cell(row=4 + i, column=1, value=line).font = normal_font()
    else:
        row = 3
        # DCF
        dcf = val.get("dcf", {})
        if dcf:
            write_section_title(ws3, row, 1, "DCF Valuation", 4)
            row += 1
            write_header_row(ws3, row, ["Metric", "Value"], 1)
            row += 1
            for k, v in dcf.items():
                write_data_row(ws3, row, [k.replace("_", " ").title(), str(v)], shade=(row % 2 == 0))
                row += 1

        row += 1
        # SOTP
        sotp = val.get("sotp", [])
        if sotp:
            write_section_title(ws3, row, 1, "Sum-of-Parts (SOTP)", 4)
            row += 1
            if isinstance(sotp, list) and sotp:
                headers = list(sotp[0].keys()) if isinstance(sotp[0], dict) else ["Value"]
                write_header_row(ws3, row, headers, 1)
                row += 1
                for i, item in enumerate(sotp):
                    if isinstance(item, dict):
                        write_data_row(ws3, row, list(item.values()), shade=(i % 2 == 0))
                    else:
                        write_data_row(ws3, row, [str(item)])
                    row += 1
            else:
                write_data_row(ws3, row, [str(sotp)])
                row += 1

        row += 1
        # Trading comps
        comps = val.get("trading_comps", [])
        if comps:
            write_section_title(ws3, row, 1, "Trading Comparables", 4)
            row += 1
            if isinstance(comps, list) and comps:
                headers = list(comps[0].keys()) if isinstance(comps[0], dict) else ["Value"]
                write_header_row(ws3, row, headers, 1)
                row += 1
                for i, item in enumerate(comps):
                    if isinstance(item, dict):
                        write_data_row(ws3, row, list(item.values()), shade=(i % 2 == 0))
                    row += 1
            else:
                write_data_row(ws3, row, [str(comps)])
                row += 1

        row += 1
        # Recommended range
        rec = val.get("recommended_ev_range", "")
        pos = val.get("positioning_rationale", "")
        if rec or pos:
            write_section_title(ws3, row, 1, "Positioning", 4)
            row += 1
            write_data_row(ws3, row, ["Recommended EV Range", str(rec)])
            row += 1
            ws3.merge_cells(start_row=row, start_column=2, end_row=row + 3, end_column=4)
            ws3.cell(row=row, column=1, value="Positioning Rationale").font = Font(bold=True, name="Calibri", size=10)
            c = ws3.cell(row=row, column=2, value=str(pos))
            c.font = normal_font()
            c.alignment = Alignment(wrap_text=True, vertical="top")
            row += 4

        row += 1
        # Precedent Transactions
        prec = val.get("precedent_transactions", [])
        if prec:
            write_section_title(ws3, row, 1, "Precedent Transactions (Indian Power M&A)", 4)
            row += 1
            if isinstance(prec, list) and prec and isinstance(prec[0], dict):
                headers = list(prec[0].keys())
                write_header_row(ws3, row, headers, 1)
                row += 1
                for i, item in enumerate(prec):
                    write_data_row(ws3, row, list(item.values()), shade=(i % 2 == 0))
                    row += 1
            row += 1

        # Football Field — structured
        ff = val.get("football_field", {})
        if ff:
            write_section_title(ws3, row, 1, "Football Field — Valuation Range (INR Cr EV)", 4)
            row += 1
            write_header_row(ws3, row, ["Methodology", "Low", "Mid", "High"], 1)
            row += 1
            for method, vals in ff.items():
                if isinstance(vals, dict):
                    write_data_row(ws3, row, [
                        method.replace("_", " ").title(),
                        vals.get("low_cr", vals.get("low", "-")),
                        vals.get("mid_cr", vals.get("mid", "-")),
                        vals.get("high_cr", vals.get("high", "-")),
                    ], shade=(row % 2 == 0))
                else:
                    write_data_row(ws3, row, [method.replace("_", " ").title(), str(vals), "", ""], shade=(row % 2 == 0))
                row += 1
            row += 1

        # Any remaining fields
        skip = {"dcf", "sotp", "trading_comps", "football_field", "precedent_transactions", "recommended_ev_range", "positioning_rationale"}
        remaining = {k: v for k, v in val.items() if k not in skip}
        if remaining:
            write_section_title(ws3, row, 1, "Additional Valuation Data", 4)
            row += 1
            for k, v in flatten(remaining):
                write_data_row(ws3, row, [k, str(v)], shade=(row % 2 == 0))
                row += 1

    for col, w in [(1, 30), (2, 22), (3, 22), (4, 22), (5, 20), (6, 20)]:
        set_col_width(ws3, col, w)

    # ── Formula Reference Table (Valuation) ─────────────────────
    row += 2
    write_section_title(ws3, row, 1, "Formula Reference", 3)
    row += 1
    write_header_row(ws3, row, ["Output", "Formula", "Notes"], 1)
    row += 1
    val_formulas = [
        ("Cost of Equity (CAPM)",  "Re = Rf + β × (Rm − Rf)",                              "Rf = risk-free rate, β = beta, Rm = market return"),
        ("WACC",                   "WACC = (E/V) × Re + (D/V) × Rd × (1 − Tax Rate)",      "E = equity, D = debt, V = E+D"),
        ("Discounted FCF",         "PV = FCF_t / (1 + WACC)^t",                             "Sum over projection period"),
        ("Terminal Value",         "TV = FCF_last × (1 + g) / (WACC − g)",                  "Gordon Growth Model; g = terminal growth rate"),
        ("Enterprise Value (DCF)", "EV = Σ PV(FCF) + PV(Terminal Value)",                   "Total DCF-based EV"),
        ("Equity Value",           "Equity Value = EV − Net Debt",                           "EV bridge to equity"),
        ("Implied Price",          "Price = Equity Value / Shares Outstanding",               "Per share value"),
        ("Upside / Downside",      "Upside % = (Implied Price − Current Price) / Current Price × 100", "vs current market price"),
        ("SOTP Segment EV",        "Segment EV = Segment EBITDA × EV/EBITDA Multiple",       "Sum of parts per segment"),
        ("Total SOTP EV",          "SOTP EV = Σ Segment EV",                                 "Aggregated across all segments"),
        ("Trading Comps EV",       "EV = EBITDA × Peer Median EV/EBITDA",                    "Market-derived multiple"),
    ]
    for i, (out, formula, note) in enumerate(val_formulas):
        write_data_row(ws3, row, [out, formula, note], shade=(i % 2 == 0))
        row += 1

    # ══════════════════════════════════════════════════════════════
    # Sheet 4: BENCHMARKING
    # ══════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Benchmarking")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:F1")
    t4 = ws4["A1"]
    t4.value = "PEER BENCHMARKING — OPERATIONAL & FINANCIAL METRICS"
    t4.font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    t4.fill = fill(DARK_BLUE)
    t4.alignment = center()
    ws4.row_dimensions[1].height = 35

    bm = final_state.get("benchmarking") or {}

    if bm.get("parse_error"):
        ws4.cell(row=3, column=1, value="Raw Output:").font = Font(bold=True, name="Calibri")
        for i, line in enumerate(str(bm.get("raw_output", "")).split("\n")[:200]):
            ws4.cell(row=4 + i, column=1, value=line).font = normal_font()
    else:
        row = 3
        # Peers
        peers = bm.get("peers", [])
        if peers:
            write_section_title(ws4, row, 1, "Peer Universe", 4)
            row += 1
            if isinstance(peers, list) and peers and isinstance(peers[0], dict):
                headers = list(peers[0].keys())
                write_header_row(ws4, row, headers, 1)
                row += 1
                for i, p in enumerate(peers):
                    write_data_row(ws4, row, list(p.values()), shade=(i % 2 == 0))
                    row += 1
            else:
                for p in peers:
                    write_data_row(ws4, row, [str(p)])
                    row += 1

        row += 1
        # Financial benchmarking
        fin_bm = bm.get("financial_benchmarking", [])
        if fin_bm:
            write_section_title(ws4, row, 1, "Financial Benchmarking", 5)
            row += 1
            if isinstance(fin_bm, list) and fin_bm and isinstance(fin_bm[0], dict):
                headers = list(fin_bm[0].keys())
                write_header_row(ws4, row, headers, 1)
                row += 1
                for i, item in enumerate(fin_bm):
                    write_data_row(ws4, row, list(item.values()), shade=(i % 2 == 0))
                    row += 1
            else:
                flat = flatten(fin_bm)
                write_header_row(ws4, row, ["Metric", "Value"], 1)
                row += 1
                for k, v in flat:
                    write_data_row(ws4, row, [k, v], shade=(row % 2 == 0))
                    row += 1

        row += 1
        # Operational benchmarking
        op_bm = bm.get("operational_benchmarking", [])
        if op_bm:
            write_section_title(ws4, row, 1, "Operational Benchmarking", 5)
            row += 1
            if isinstance(op_bm, list) and op_bm and isinstance(op_bm[0], dict):
                headers = list(op_bm[0].keys())
                write_header_row(ws4, row, headers, 1)
                row += 1
                for i, item in enumerate(op_bm):
                    write_data_row(ws4, row, list(item.values()), shade=(i % 2 == 0))
                    row += 1

        row += 1
        # Key takeaways
        takeaways = bm.get("key_takeaways", [])
        if takeaways:
            write_section_title(ws4, row, 1, "Key Takeaways", 5)
            row += 1
            for i, t in enumerate(takeaways):
                ws4.cell(row=row, column=1, value=f"• {t}").font = normal_font()
                ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                ws4.cell(row=row, column=1).fill = fill(LIGHT_GREY if i % 2 == 0 else WHITE)
                row += 1

        row += 1
        # Outliers
        outliers = bm.get("outliers", [])
        if outliers:
            write_section_title(ws4, row, 1, "Outlier Flags", 5)
            row += 1
            if isinstance(outliers, list) and outliers and isinstance(outliers[0], dict):
                headers = list(outliers[0].keys())
                write_header_row(ws4, row, headers, 1)
                row += 1
                for i, item in enumerate(outliers):
                    write_data_row(ws4, row, list(item.values()), shade=(i % 2 == 0))
                    row += 1

    for col, w in [(1, 30), (2, 20), (3, 20), (4, 20), (5, 20)]:
        set_col_width(ws4, col, w)

    # ── Formula Reference Table (Benchmarking) ───────────────────
    row += 2
    write_section_title(ws4, row, 1, "Formula Reference", 3)
    row += 1
    write_header_row(ws4, row, ["Metric", "Formula", "Notes"], 1)
    row += 1
    bm_formulas = [
        ("EV / EBITDA",         "EV / EBITDA",                                              "Primary valuation multiple for capital-intensive sectors"),
        ("P/E Ratio",           "Market Price per Share / Earnings per Share (EPS)",         "Equity market multiple"),
        ("Net Debt / EBITDA",   "Net Debt / EBITDA",                                         "Leverage ratio; lower = stronger balance sheet"),
        ("ROE",                 "Net Profit / Shareholders Equity × 100",                    "Return on equity (%)"),
        ("EBITDA Margin",       "EBITDA / Revenue × 100",                                    "Operating profitability (%)"),
        ("Revenue CAGR",        "(Revenue_end / Revenue_start) ^ (1/n) − 1",                 "Compound annual growth rate over n years"),
        ("EV / MW",             "Enterprise Value / Installed Capacity (MW)",                "Capacity-based valuation for power sector"),
        ("Peer Median",         "Median of all peer values for each metric",                  "Used as benchmark vs target company"),
    ]
    for i, (metric, formula, note) in enumerate(bm_formulas):
        write_data_row(ws4, row, [metric, formula, note], shade=(i % 2 == 0))
        row += 1

    # ══════════════════════════════════════════════════════════════
    # Sheet 5: ASSEMBLY / REVIEW
    # ══════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("Assembly / Review")
    ws7.sheet_view.showGridLines = False

    ws7.merge_cells("A1:D1")
    t7 = ws7["A1"]
    t7.value = "ASSEMBLY / REVIEW — QC REPORT"
    t7.font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    t7.fill = fill(DARK_BLUE)
    t7.alignment = center()
    ws7.row_dimensions[1].height = 35

    cr = final_state.get("consistency_report") or {}
    row = 3

    write_section_title(ws7, row, 1, "Overall Status", 3)
    row += 1
    status_val = cr.get("status", "N/A").upper()
    status_color = "00B050" if "PASS" in status_val else "FF0000"
    c = ws7.cell(row=row, column=1, value=f"Status: {status_val}")
    c.font = Font(name="Calibri", bold=True, color=status_color, size=12)
    ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    issues = cr.get("issues", [])
    if issues:
        write_section_title(ws7, row, 1, f"Issues Found ({len(issues)})", 3)
        row += 1
        write_header_row(ws7, row, ["#", "Issue"], 1)
        ws7.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1
        for i, issue in enumerate(issues):
            txt = issue if isinstance(issue, str) else issue.get("description", str(issue))
            ws7.cell(row=row, column=1, value=i + 1).font = normal_font()
            c = ws7.cell(row=row, column=2, value=txt)
            c.font = normal_font()
            c.alignment = Alignment(wrap_text=True)
            ws7.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            if i % 2 == 0:
                ws7.cell(row=row, column=1).fill = fill(LIGHT_GREY)
                ws7.cell(row=row, column=2).fill = fill(LIGHT_GREY)
            row += 1

    row += 1
    recs = cr.get("recommendations", [])
    if recs:
        write_section_title(ws7, row, 1, f"Recommendations ({len(recs)})", 3)
        row += 1
        write_header_row(ws7, row, ["#", "Recommendation"], 1)
        ws7.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1
        for i, rec in enumerate(recs):
            txt = rec if isinstance(rec, str) else str(rec)
            ws7.cell(row=row, column=1, value=i + 1).font = normal_font()
            c = ws7.cell(row=row, column=2, value=txt)
            c.font = normal_font()
            c.alignment = Alignment(wrap_text=True)
            ws7.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            if i % 2 == 0:
                ws7.cell(row=row, column=1).fill = fill(LIGHT_GREY)
                ws7.cell(row=row, column=2).fill = fill(LIGHT_GREY)
            row += 1

    for col, w in [(1, 6), (2, 55), (3, 30)]:
        set_col_width(ws7, col, w)

    # ── Save workbook ──────────────────────────────────────────
    wb.save(output_path)
    print(f"\n  ✅ Excel saved → {output_path}")




# ─────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import os
    from dotenv import load_dotenv
    from datetime import datetime

    load_dotenv()

    final_state, feedback = run_analyst_pipeline()

    output_dir = os.path.join(os.path.dirname(__file__), "outputs")
    os.makedirs(output_dir, exist_ok=True)
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    ctx        = final_state.get("deal_context") or {}
    safe_name  = ctx.get("target_name", "Company").replace(" ", "_")

    excel_path = os.path.join(output_dir, f"{safe_name}_{timestamp}.xlsx")

    write_excel(final_state, feedback, excel_path)

    cr = final_state.get("consistency_report") or {}
    print(f"\n{'='*60}")
    print("PIPELINE COMPLETE")
    print(f"{'='*60}")
    print(f"  Decision   : {feedback['decision'].upper()}")
    print(f"  QC Status  : {cr.get('status', 'N/A').upper()}")
    print(f"  Excel      : {excel_path}")
    print(f"{'='*60}")
